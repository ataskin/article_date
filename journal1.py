import time
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from selenium import webdriver

wb = load_workbook("links1.xlsx") # set the excel file name and sheet name
ws = wb["Sheet1"]
# please Input number of ecxel row below
excel= 411
for row in range(1,excel+1):
    url = ws.cell(int(row),1).value # get link from excel
    browser=webdriver.Firefox() # start web browser -> Firefox

    browser.get(url) # get url content
    time.sleep(7)  # wait 7 seconds
    html = browser.page_source 
    soup = BeautifulSoup(html, "html.parser")
    source = str(soup.encode('utf-8')) # encode as utf-8
    field = ''
    date = re.search('"dates":(.*),"displayViewFullText":', source)
    date = date.group(1)
    dates = date.replace('{', '')
    dates = dates.replace('}', '')
    dates = dates.replace('"', '')

    get_source = (str(dates))
    hs = open("journal1.txt","a")
    hs.write(url +'; ' + get_source + "\n") # write url + finding as new line to the txt file
    hs.close()
    browser.close() # close web browser
    print (row, 'OK')
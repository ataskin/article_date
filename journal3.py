import requests
import pandas as pd
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook

wb = load_workbook("links3.xlsx") # set the excel file name and sheet name
ws = wb["Sheet1"]
# please Input number of ecxel row below
excel= 1793

for row in range(1,excel+1):
    url = (ws.cell(int(row),1).value) # get link from excel
    r  = requests.get(url)
    time.sleep(2) # wait 2 seconds
    data = r.text
    time.sleep(3) # wait 3 seconds
    soup = BeautifulSoup(data, "html.parser")

    need = [] # define list variable
    # get content in <ul> and  <li> tags
    for ul_tag in soup.find_all('ul', {'class':'c-bibliographic-information__list'}): 
        for li_tag in ul_tag.find_all('li'):
            field = li_tag.find('p').text
            need.append(field) # add the findings to the list

    if need ==[]: # if the list is empty write the url 
        hs = open("journal3.txt","a")
        hs.write(url +'; ' + str(need) + "\n")
        hs.close()

        print(row, 'Empty --->',url)
    else:   # if the list is full write the url and findings 
        hs = open("journal3.txt","a")
        hs.write(url +'; ' + str(need) + "\n")
        hs.close()
        print(row, 'OK -->',url)

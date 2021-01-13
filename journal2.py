import time
from bs4 import BeautifulSoup
from openpyxl import Workbook,load_workbook
from selenium import webdriver

wb = load_workbook("links2.xlsx") # set the excel file name and sheet name
ws = wb["Sheet1"]
# please Input number of ecxel row below
excel= 238
for row in range(1,excel+1):
    url = ws.cell(int(row),1).value # get link from excel
    browser=webdriver.Firefox()# start web browser -> Firefox
    time.sleep(1) # wait 1 second
    browser.get(url) # get url content
    time.sleep(7) # wait 7 seconds
    html = browser.page_source 
    time.sleep(1) # wait 1 second
    soup = BeautifulSoup(html, "html.parser")
    time.sleep(1) # wait 1 second
    field = ''
    # get content in <section> and  <ul> tags
    for section_tag in soup.find_all('section', {'class':'publication-history'}):
        field = section_tag.find('ul').text 
    field = field.replace('\n\n', '\n') 
    need = field.replace('\n',';')
    
    hs = open("journal2.txt","a")
    hs.write(url +'; ' + need + "\n") # write url + finding as new line to the txt file
    hs.close() 

    browser.close() # close web browser
    print (row, url, ' OK')